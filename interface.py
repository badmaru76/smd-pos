import cv2
import matplotlib.pyplot as plt
import sys
from openpyxl import load_workbook
import pandas as pd


sys.path.append('/home/badmaru/Documenti')

smd_df = pd.DataFrame()
x_scale = 1
y_scale = 1
x_origin = 0
y_origin = 0
x_mouse = 0
y_mouse = 0
click_origin = 0
x_end = 0
y_end = 0
x_dim = 0
y_dim = 0



def dbClick_event(event,x ,y , flags, param):
    global x_mouse
    global y_mouse

    if event == cv2.EVENT_LBUTTONDBLCLK:
        x_mouse, y_mouse = x, y


def set_origin(x, y):
    global x_origin
    global y_origin

    #cv2.circle(img, (x, y), 3, (255, 0, 0), -1)
    x_origin = x
    y_origin = y


def set_scale(x_mm, y_mm ,x ,y ):
    global x_scale
    global y_scale
    global x_origin
    global y_origin
    global x_end
    global y_end

    #cv2.circle(img, (x, y), 3, (255, 0, 0), -1)
    x_end = x
    y_end = y
    x_scale = (x_end - x_origin) / x_mm
    y_scale = (y_origin - y_end) / y_mm


def calc_position(x, y, rot):
    global x_scale
    global y_scale

    x = x_origin + (x * x_scale)
    y = y_origin - (y * y_scale)
    rot = rot
    return x, y, rot


def get_position(x_mouse, y_mouse, rot):
    global x_scale, x_origin
    global y_scale, y_origin

    x_mm = ((x_mouse - x_origin) / x_scale)
    y_mm = ((y_origin - y_mouse) / y_scale)
    rot = rot
    update_PP('U1', x_mm, y_mm)
    return x_mm, y_mm, rot


def load_datafile(filename):
    global x_dim, y_dim, smd_df
    itemcode = 0
    ef = 0
    wb = load_workbook(filename)
    #ws = wb["Sheet1"]
    ws = wb.active

    #Set PCB dimension
    x_dim = float(ws['C1'].value)
    y_dim = float(ws['D1'].value)
    #rot = ws['D2'].value
    #designator = ws['A2'].value
    row_count = ws.max_row
    smd_df = pd.DataFrame(columns=['ItemCode', 'Ref', 'X', 'Y', 'Rot'])

    for row_cells in ws.iter_rows(min_row=2, max_row=row_count):
        itemcode = row_cells[0].value
        ref = row_cells[1].value
        pos_x = row_cells[2].value
        pos_y = row_cells[3].value
        rot = row_cells[4].value
        row = {'ItemCode':itemcode, 'Ref':ref, 'X':pos_x, 'Y':pos_y,'Rot':rot}
        new_df = pd.DataFrame([row])
        smd_df = pd.concat([smd_df, new_df], axis=0, ignore_index=True)
    return smd_df


def update_PP(ref, pos_x_mm, pos_y_mm):
    global smd_df
    smd_df.loc[smd_df['Ref'] == ref, 'X'] = pos_x_mm
    smd_df.loc[smd_df['Ref'] == ref, 'Y'] = pos_y_mm
    print(smd_df.loc[smd_df['Ref'] == ref])

